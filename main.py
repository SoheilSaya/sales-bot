import logging
from telegram import Update, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import Application, MessageHandler, filters, CommandHandler
from openpyxl import Workbook, load_workbook
from datetime import datetime
import requests
# Define your token
TOKEN = "7876541WxedE1iaH_dXg"

# Define the manager's ID
MANAGER_ID = 1113600227

# Define the keywords you are looking for
keywords = ["قیمت", "موجودی", "میلگرد", "خرید", "ورق", "تیراهن"]

# Initialize Excel files
messages_wb = Workbook()
messages_ws = messages_wb.active
messages_ws.append(["Keyword", "Message", "Sender ID", "Sender Name", "Telegram Username", "Phone Number", "Date Time", "Message Link"])

phonebook_wb = Workbook()
phonebook_ws = phonebook_wb.active
phonebook_ws.append(["User ID", "Telegram Username", "Phone Number"])

# Setup logging
logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                    level=logging.INFO)
logger = logging.getLogger(__name__)

def save_to_excel(keyword, message, sender_id, sender_name, username, phone_number, message_link):
    """Saves relevant data to the messages Excel file"""
    date_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    messages_ws.append([keyword, message, sender_id, sender_name, username, phone_number, date_time, message_link])
    messages_wb.save("messages.xlsx")

def save_phone_number(user_id, username, phone_number):
    """Saves phone number and username to the phonebook Excel file"""
    phonebook_ws.append([user_id, username, phone_number])
    phonebook_wb.save("phonebook.xlsx")

def check_duplicate_number(phone_number):
    """Checks if a phone number already exists in the phonebook"""
    try:
        phonebook_wb = load_workbook("phonebook.xlsx")
        phonebook_ws = phonebook_wb.active
        for row in phonebook_ws.iter_rows(min_row=2, values_only=True):
            if row[2] == phone_number:
                return True
    except FileNotFoundError:
        pass  # If phonebook doesn't exist yet, no problem
    return False

async def message_handler(update: Update, context):
    """Handles incoming messages and checks for keywords"""
    message = update.message.text
    sender_id = update.message.from_user.id
    sender_name = update.message.from_user.full_name
    username = f"@{update.message.from_user.username}" if update.message.from_user.username else "N/A"

    # Look up the phone number in the phonebook
    phone_number = "N/A"
    try:
        phonebook_wb = load_workbook("phonebook.xlsx")
        phonebook_ws = phonebook_wb.active
        for row in phonebook_ws.iter_rows(min_row=2, values_only=True):
            if row[0] == sender_id:
                phone_number = row[2]
                break
    except FileNotFoundError:
        pass  # If phonebook doesn't exist yet, no problem, phone_number remains "N/A"

    matched_keywords = []
    for keyword in keywords:
        if keyword in message:
            matched_keywords.append(keyword)

    # If any keywords match, save message to Excel and notify the manager
    if matched_keywords:
        # Create a link to the message in Telegram (using message URL format)
        message_link = f"https://t.me/{update.message.chat.username}/{update.message.message_id}"
        for keyword in matched_keywords:
            save_to_excel(keyword, message, sender_id, sender_name, username, phone_number, message_link)
        
        # Notify the manager about the keyword detection
        notification = (
            f"🔔 **هشدار کلمه کلیدی**\n\n"
            f"📌 **کلمات کلیدی**: {', '.join(matched_keywords)}\n"
            f"👤 **نام فرستنده**: {sender_name}\n"
            f"🆔 **آیدی**: {sender_id}\n"
            f"🌐 **یوزرنیم**: {username}\n"
            f"📨 **متن پیام**: {message}\n"
            f"📞 **شماره موبایل**: {phone_number}\n"
            f"🔗 **لینک پیام**: {message_link}"
        )
        await context.bot.send_message(chat_id=MANAGER_ID, text=notification, parse_mode="Markdown")

async def start(update: Update, context):
    """Start command to send buttons"""
    user = update.message.from_user
    user_id = user.id

    # Check if phone number is already in phonebook
    try:
        phonebook_wb = load_workbook("phonebook.xlsx")
        phonebook_ws = phonebook_wb.active
        phone_number_exists = any(row[0] == user_id for row in phonebook_ws.iter_rows(min_row=2, values_only=True))
    except FileNotFoundError:
        phone_number_exists = False

    # Buttons to share phone number and check if number is saved
    keyboard = [
        [KeyboardButton("اشتراک گذاری شماره 📲", request_contact=True)]
    ]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    
    if not phone_number_exists:
        await update.message.reply_text(
            "سلام لطفا شماره خود را شیر کنید.",
            reply_markup=reply_markup
        )
    else:
        await update.message.reply_text(
            "شماره شما ذخیره شده است.",
            reply_markup=reply_markup
        )

async def contact_handler(update: Update, context):
    """Handles the phone number shared by the user"""
    user = update.message.from_user
    user_id = user.id
    username = f"@{user.username}" if user.username else "N/A"
    phone_number = update.message.contact.phone_number

    # Check if the phone number is already in the phonebook
    if check_duplicate_number(phone_number):
        await update.message.reply_text(f"شماره شما {phone_number} قبلا ذخیره شده است.")
    else:
        # Save phone number and username to the phonebook
        save_phone_number(user_id, username, phone_number)
        await update.message.reply_text(f"مرسی، شماره شما ({phone_number}) ذخیره شد.")

def main():
    """Start the bot"""
    application = Application.builder().token(TOKEN).build()

    # Commands
    application.add_handler(CommandHandler("start", start))

    # Message handlers
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, message_handler))

    # Contact handler
    application.add_handler(MessageHandler(filters.CONTACT, contact_handler))

    # Start the bot
    application.run_polling()

if __name__ == '__main__':
    main()
